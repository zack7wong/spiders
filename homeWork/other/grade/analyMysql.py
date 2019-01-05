#!/usr/bin/env python
# -*- coding:utf-8 -*-

import execMysql
from openpyxl import load_workbook

class AnalyMysql():
    def __init__(self):
        self.exe = execMysql.ExecMysql()
        print('已实例化数据库操作类')

    def average(self,item):
        myavg = float("%.2f" % ((item['liucheng'] + item['uml'] + item['php']) / 3))
        item['avg'] = myavg

    def total(self,item):
        mytotal = float("%.2f" % (item['liucheng'] + item['uml'] + item['php']))
        item['total'] = mytotal

    def por_total(self,item):
        zongping = float("%.2f" % (item['liucheng'] * 0.15 + item['uml'] * 0.25 + item['php'] * 0.6))
        item['zongping'] = zongping

    def rank(self,item_list,mytype=''):

        if mytype == 'rank':
            all_item = sorted(item_list, key=lambda x: x['avg'], reverse=True)
        else:
            all_item = sorted(item_list, key=lambda x: x['zongping'], reverse=True)

        num = 0
        temp = ''
        myflag = 1
        for item in all_item:
            if item['avg'] != temp:
                num = myflag
                temp = item['avg']
            if mytype =='rank':
                item['rank'] = num
            else:
                item['zongpingRank'] = num
            myflag +=1

    def grade(self,myitem):
        grade_dic = [{"key": [0, 59], "value": "不及格"}, {"key": [60, 69], "value": "及格"},{"key": [70, 84], "value": "良好"}, {"key": [85, 100], "value": "优秀"}]
        for item in grade_dic:
            if myitem['avg'] >= item['key'][0] and myitem['avg'] <=item['key'][1]:
                myitem['gradeSore'] = item['value']


if __name__ == '__main__':
    ana = AnalyMysql()

    wb = load_workbook('python数据.xlsx')
    sheet = wb.active
    # print(sheet.max_row)
    item_list = []
    for i in range(1,sheet.max_row):
        hang_list = []
        for cell in list(sheet.rows)[i]:
            hang_list.append(cell.value)
        obj = {
            'name':hang_list[0],
            'liucheng':hang_list[1],
            'uml':hang_list[2],
            'php':hang_list[3],
            'avg':hang_list[4],
            'rank':hang_list[5],
            'gradeSore':hang_list[6],
        }
        item_list.append(obj)

    for item in item_list:
        ana.average(item)
        ana.total(item)
        ana.grade(item)
        ana.por_total(item)

    ana.rank(item_list,'rank')
    ana.rank(item_list)
    print('数据插入完成')
    for item in item_list:
        ana.exe.insert(item['name'],item['liucheng'],item['uml'],item['php'],item['avg'],item['rank'],item['gradeSore'])
        print('姓名：'+item['name']+' 平均分为：'+str(item['avg'])+' 等级为：'+item['gradeSore']+' 总分为：'+str(item['total'])+' 总分排名:'+str(item['rank'])+' 总评成绩：'+str(item['zongping'])+' 总评排名：'+str(item['zongpingRank']) )

    for item in item_list:
        if item['name'] == '姚铜就' or item['name'] == '张子额':
            print('成功查找：'+item['name']+' 平均分为：'+str(item['avg'])+' 等级为：'+item['gradeSore']+' 总分为：'+str(item['total'])+' 总分排名:'+str(item['rank'])+' 总评成绩：'+str(item['zongping'])+' 总评排名：'+str(item['zongpingRank']) )

    print('不及格的同学有:')
    for item in item_list:
        if item['avg'] < 60 or item['zongping'] <60:
            print(item['name'])